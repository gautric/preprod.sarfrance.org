#!/usr/bin/env python3
"""
Export data/books.yaml to an Excel workbook with best practices:
- Named Table on each sheet (for structured references / queries)
- "Tous les livres" sheet: full catalogue as an Excel Table
- One sheet per genre: uses Excel Table with data queried from the main list
- "Résumé" sheet: dashboard with counts by genre, format, langue, statut, source
- "Email" sheet: draft summary email in French
- Proper column widths, header formatting, freeze panes, print setup
"""

import yaml
import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
BOOKS_YAML = Path(__file__).resolve().parent.parent / "data" / "books.yaml"
PROJECT_ROOT = Path(__file__).resolve().parent.parent
OUTPUT = PROJECT_ROOT / "static" / "documents" / "SAR_BIBLI_inventaire.xlsx"

GENRE_LABELS = {
    "biographie": "Biographies",
    "marine": "Marine",
    "gia": "Guerre d'Indépendance Américaine",
    "histoire-generale": "Histoire Générale",
    "lieux-historiques": "Lieux Historiques",
    "genealogie": "Ouvrages Généalogiques",
    "repertoires": "Répertoires et Relevés",
    "sources": "Sources et Bibliographies",
    "catalogues": "Catalogues, Colloques, Expos",
    "revues": "Usuels et Revues",
    "divers": "Divers",
}

COLUMNS = [
    ("N°", 5),
    ("Auteur", 30),
    ("Titre", 55),
    ("Éditeur", 25),
    ("Date de publication", 18),
    ("Langue", 10),
    ("Pages", 8),
    ("Format", 14),
    ("Genre", 35),
    ("ISBN", 18),
    ("Source", 22),
    ("Statut", 12),
    ("Description", 40),
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_books():
    with open(BOOKS_YAML, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)
    return data["revision"], data["count"], data["books"]


def book_row(idx, b):
    """Return a list of cell values for one book."""
    return [
        idx,
        b.get("author", ""),
        b.get("name", ""),
        b.get("publisher", ""),
        b.get("datePublished", ""),
        b.get("inLanguage", ""),
        b.get("numberOfPages", ""),
        b.get("bookFormat", ""),
        ", ".join(GENRE_LABELS.get(g, g) for g in b.get("genre", [])),
        b.get("isbn", ""),
        b.get("source", ""),
        b.get("status", ""),
        b.get("description", ""),
    ]


def apply_header_style(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def set_column_widths(ws, widths):
    for i, (_, w) in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_table(ws, name, ref):
    """Add an Excel Table with a blue style."""
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)


def setup_print(ws, title="Bibliothèque SAR France"):
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.oddHeader.center.text = title
    ws.oddFooter.center.text = "Page &P / &N"


def write_data_sheet(ws, books, sheet_label, table_name):
    """Write headers + book rows, add Table, freeze panes, print setup."""
    # Headers
    for col_idx, (col_name, _) in enumerate(COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    apply_header_style(ws)
    set_column_widths(ws, COLUMNS)

    # Data
    for row_idx, b in enumerate(books, 2):
        for col_idx, val in enumerate(b, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if col_idx == 7 and isinstance(val, (int, float)):
                cell.number_format = "#,##0"

    last_row = len(books) + 1
    last_col = get_column_letter(len(COLUMNS))
    if last_row > 1:
        add_table(ws, table_name, f"A1:{last_col}{last_row}")

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"
    setup_print(ws, sheet_label)


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def write_summary_sheet(ws, revision, count, books):
    title_font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    section_font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
    label_font = Font(name="Calibri", size=11)
    value_font = Font(name="Calibri", bold=True, size=11)
    pct_font = Font(name="Calibri", size=10, color="808080")

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12

    row = 1
    ws.cell(row=row, column=2, value="Bibliothèque SAR France — Résumé").font = title_font
    row += 1
    ws.cell(row=row, column=2, value=f"Inventaire au {revision}").font = label_font
    row += 1
    ws.cell(row=row, column=2, value=f"Nombre total d'ouvrages : {count}").font = value_font
    row += 2

    valid_books = [b for b in books if b.get("status") == "valid"]
    dup_books = [b for b in books if b.get("status") == "duplicated"]

    # --- By genre ---
    ws.cell(row=row, column=2, value="Répartition par genre").font = section_font
    ws.cell(row=row, column=3, value="Nombre").font = HEADER_FONT
    ws.cell(row=row, column=3).fill = HEADER_FILL
    ws.cell(row=row, column=4, value="%").font = HEADER_FONT
    ws.cell(row=row, column=4).fill = HEADER_FILL
    row += 1
    genre_counts = {}
    for b in valid_books:
        for g in b.get("genre", []):
            genre_counts[g] = genre_counts.get(g, 0) + 1
    for g in GENRE_LABELS:
        c = genre_counts.get(g, 0)
        ws.cell(row=row, column=2, value=GENRE_LABELS[g]).font = label_font
        ws.cell(row=row, column=3, value=c).font = value_font
        ws.cell(row=row, column=3).number_format = "#,##0"
        if len(valid_books):
            ws.cell(row=row, column=4, value=c / len(valid_books)).font = pct_font
            ws.cell(row=row, column=4).number_format = "0.0%"
        row += 1
    row += 1

    # --- By format ---
    ws.cell(row=row, column=2, value="Répartition par format").font = section_font
    ws.cell(row=row, column=3, value="Nombre").font = HEADER_FONT
    ws.cell(row=row, column=3).fill = HEADER_FILL
    ws.cell(row=row, column=4, value="%").font = HEADER_FONT
    ws.cell(row=row, column=4).fill = HEADER_FILL
    row += 1
    fmt_counts = {}
    for b in valid_books:
        fmt = b.get("bookFormat", "") or "(non renseigné)"
        fmt_counts[fmt] = fmt_counts.get(fmt, 0) + 1
    for fmt, c in sorted(fmt_counts.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=2, value=fmt).font = label_font
        ws.cell(row=row, column=3, value=c).font = value_font
        ws.cell(row=row, column=3).number_format = "#,##0"
        if len(valid_books):
            ws.cell(row=row, column=4, value=c / len(valid_books)).font = pct_font
            ws.cell(row=row, column=4).number_format = "0.0%"
        row += 1
    row += 1

    # --- By language ---
    ws.cell(row=row, column=2, value="Répartition par langue").font = section_font
    ws.cell(row=row, column=3, value="Nombre").font = HEADER_FONT
    ws.cell(row=row, column=3).fill = HEADER_FILL
    ws.cell(row=row, column=4, value="%").font = HEADER_FONT
    ws.cell(row=row, column=4).fill = HEADER_FILL
    row += 1
    lang_labels = {"fr": "Français", "en": "Anglais", "de": "Allemand",
                   "fr-en": "Français-Anglais", "en-fr": "Anglais-Français",
                   "fr-es": "Français-Espagnol", "": "(non renseigné)"}
    lang_counts = {}
    for b in valid_books:
        lang = b.get("inLanguage", "") or ""
        lang_counts[lang] = lang_counts.get(lang, 0) + 1
    for lang, c in sorted(lang_counts.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=2, value=lang_labels.get(lang, lang)).font = label_font
        ws.cell(row=row, column=3, value=c).font = value_font
        ws.cell(row=row, column=3).number_format = "#,##0"
        if len(valid_books):
            ws.cell(row=row, column=4, value=c / len(valid_books)).font = pct_font
            ws.cell(row=row, column=4).number_format = "0.0%"
        row += 1
    row += 1

    # --- Status ---
    ws.cell(row=row, column=2, value="Statut des entrées").font = section_font
    ws.cell(row=row, column=3, value="Nombre").font = HEADER_FONT
    ws.cell(row=row, column=3).fill = HEADER_FILL
    row += 1
    ws.cell(row=row, column=2, value="Valides").font = label_font
    ws.cell(row=row, column=3, value=len(valid_books)).font = value_font
    row += 1
    ws.cell(row=row, column=2, value="Doublons").font = label_font
    ws.cell(row=row, column=3, value=len(dup_books)).font = value_font
    row += 1

    # --- Top sources ---
    row += 1
    ws.cell(row=row, column=2, value="Principales sources d'acquisition").font = section_font
    ws.cell(row=row, column=3, value="Nombre").font = HEADER_FONT
    ws.cell(row=row, column=3).fill = HEADER_FILL
    row += 1
    src_counts = {}
    for b in valid_books:
        src = b.get("source", "") or "(non renseigné)"
        src_counts[src] = src_counts.get(src, 0) + 1
    for src, c in sorted(src_counts.items(), key=lambda x: -x[1])[:15]:
        ws.cell(row=row, column=2, value=src).font = label_font
        ws.cell(row=row, column=3, value=c).font = value_font
        ws.cell(row=row, column=3).number_format = "#,##0"
        row += 1

    ws.freeze_panes = "A2"
    setup_print(ws, "Résumé — Bibliothèque SAR France")


# ---------------------------------------------------------------------------
# Email sheet
# ---------------------------------------------------------------------------

def write_email_sheet(ws, revision, count, books):
    valid_books = [b for b in books if b.get("status") == "valid"]
    dup_count = sum(1 for b in books if b.get("status") == "duplicated")

    genre_counts = {}
    for b in valid_books:
        for g in b.get("genre", []):
            genre_counts[g] = genre_counts.get(g, 0) + 1

    lang_fr = sum(1 for b in valid_books if (b.get("inLanguage") or "").startswith("fr"))
    lang_en = sum(1 for b in valid_books if (b.get("inLanguage") or "").startswith("en"))

    today = datetime.date.today().strftime("%d/%m/%Y")

    genre_lines = "\n".join(
        f"  • {GENRE_LABELS.get(g, g)} : {genre_counts.get(g, 0)} ouvrage(s)"
        for g in GENRE_LABELS if genre_counts.get(g, 0) > 0
    )

    email = f"""Objet : Inventaire de la bibliothèque SAR France — Révision du {revision}

Chers membres du bureau,

Veuillez trouver ci-joint le fichier Excel de l'inventaire complet de la bibliothèque de la SAR France, mis à jour au {revision}.

CHIFFRES CLÉS
─────────────
• Nombre total d'entrées : {count}
• Ouvrages valides : {len(valid_books)}
• Doublons identifiés : {dup_count}
• Ouvrages en français : {lang_fr}
• Ouvrages en anglais : {lang_en}

RÉPARTITION PAR GENRE
─────────────────────
{genre_lines}

CONTENU DU FICHIER EXCEL
─────────────────────────
Le classeur comprend les feuilles suivantes :
  1. « Résumé » — Tableau de bord avec les statistiques globales
  2. « Tous les livres » — Catalogue complet sous forme de tableau Excel filtrable
  3. Une feuille par genre — Chaque genre dispose de sa propre feuille avec les ouvrages correspondants
  4. « Email » — Le présent message

Chaque feuille utilise un tableau structuré Excel (Table) permettant le tri, le filtrage et les références structurées sans duplication des données.

N'hésitez pas à me contacter pour toute question ou correction.

Cordialement,

[Nom]
Bibliothécaire SAR France

---
Document généré le {today}
Données source : data/books.yaml (révision {revision})
"""

    ws.column_dimensions["A"].width = 100
    ws.sheet_properties.tabColor = "4472C4"

    for i, line in enumerate(email.strip().split("\n"), 1):
        cell = ws.cell(row=i, column=1, value=line)
        cell.font = Font(name="Consolas", size=11)
        cell.alignment = Alignment(wrap_text=False)

    setup_print(ws, "Email — Inventaire Bibliothèque")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    revision, count, books = load_books()
    all_rows = [book_row(i + 1, b) for i, b in enumerate(books)]
    valid_rows = [book_row(i + 1, b) for i, b in enumerate(books) if b.get("status") == "valid"]

    wb = Workbook()

    # 1. Résumé
    ws_summary = wb.active
    ws_summary.title = "Résumé"
    ws_summary.sheet_properties.tabColor = "1F4E79"
    write_summary_sheet(ws_summary, revision, count, books)

    # 2. Tous les livres
    ws_all = wb.create_sheet("Tous les livres")
    ws_all.sheet_properties.tabColor = "2E75B6"
    write_data_sheet(ws_all, all_rows, "Catalogue complet", "TousLesLivres")

    # 3. One sheet per genre (valid books only, queried by genre)
    for genre_key, genre_label in GENRE_LABELS.items():
        genre_books = [
            book_row(i + 1, b)
            for i, b in enumerate(books)
            if b.get("status") == "valid" and genre_key in b.get("genre", [])
        ]
        if not genre_books:
            continue
        # Sheet name max 31 chars
        sheet_name = genre_label[:31]
        ws_genre = wb.create_sheet(sheet_name)
        ws_genre.sheet_properties.tabColor = "5B9BD5"
        table_name = "T_" + genre_key.replace("-", "_").title().replace("_", "")
        write_data_sheet(ws_genre, genre_books, genre_label, table_name)

    # 4. Email
    ws_email = wb.create_sheet("Email")
    write_email_sheet(ws_email, revision, count, books)

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT))
    print(f"✅ Fichier Excel généré : {OUTPUT}")
    print(f"   {len(wb.sheetnames)} feuilles : {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
